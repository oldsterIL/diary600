#!/usr/bin/python3
# -*- coding: utf-8 -*-

import openpyxl
from openpyxl.styles import Font
import datetime as dt
import time
import locale

import zhorik.db as db
import driver.cnl24lib as cnl24lib

breakfast = 0
lunch = 0
dinner = 0


def insert_bgl_night(start_time, db_connect, sheet):

    font_red = Font(name='Calibri', size=14, bold=True, color='FF0000')

    row = 29
    # Ночные Сахара
    bgl, time_bgl = db_connect.get_last_bgl(start_time)
    if bgl == "--" or bgl < 4 or bgl > 10:
        sheet.cell(row=row, column=3).font = font_red
    sheet.cell(row=row, column=3).value = bgl
    sheet.cell(row=row, column=2).value = time_bgl

    bgl, time_bgl = db_connect.get_last_bgl(start_time + 60 * 60)
    if bgl == "--" or bgl < 4 or bgl > 10:
        sheet.cell(row=row+1, column=3).font = font_red
    sheet.cell(row=row+1, column=3).value = bgl
    sheet.cell(row=row+1, column=2).value = time_bgl

    bgl, time_bgl = db_connect.get_last_bgl(start_time + 120 * 60)
    if bgl == "--" or bgl < 4 or bgl > 10:
        sheet.cell(row=row+2, column=3).font = font_red
    sheet.cell(row=row+2, column=3).value = bgl
    sheet.cell(row=row+2, column=2).value = time_bgl

    bgl, time_bgl = db_connect.get_last_bgl(start_time + 180 * 60)
    if bgl == "--" or bgl < 4 or bgl > 10:
        sheet.cell(row=row, column=5).font = font_red
    sheet.cell(row=row, column=5).value = bgl
    sheet.cell(row=row, column=4).value = time_bgl

    bgl, time_bgl = db_connect.get_last_bgl(start_time + 240 * 60)
    if bgl == "--" or bgl < 4 or bgl > 10:
        sheet.cell(row=row+1, column=5).font = font_red
    sheet.cell(row=row+1, column=5).value = bgl
    sheet.cell(row=row+1, column=4).value = time_bgl

    bgl, time_bgl = db_connect.get_last_bgl(start_time + 300 * 60)
    if bgl == "--" or bgl < 4 or bgl > 10:
        sheet.cell(row=row+2, column=5).font = font_red
    sheet.cell(row=row + 2, column=5).value = bgl
    sheet.cell(row=row+2, column=4).value = time_bgl

    bgl, time_bgl = db_connect.get_last_bgl(start_time + 360 * 60)
    if bgl == "--" or bgl < 4 or bgl > 10:
        sheet.cell(row=row, column=7).font = font_red
    sheet.cell(row=row, column=7).value = bgl
    sheet.cell(row=row, column=6).value = time_bgl

    bgl, time_bgl = db_connect.get_last_bgl(start_time + 420 * 60)
    if bgl == "--" or bgl < 4 or bgl > 10:
        sheet.cell(row=row+1, column=7).font = font_red
    sheet.cell(row=row + 1, column=7).value = bgl
    sheet.cell(row=row+1, column=6).value = time_bgl

    bgl, time_bgl = db_connect.get_last_bgl(start_time + 480 * 60)
    if bgl == "--" or bgl < 4 or bgl > 10:
        sheet.cell(row=row+2, column=7).font = font_red
    sheet.cell(row=row + 2, column=7).value = bgl
    sheet.cell(row=row+2, column=6).value = time_bgl


def insert_bgl(row, start_time, db_connect, sheet):

    font_red = Font(name='Calibri', size=14, bold=True, color='FF0000')

    # Сахара
    bgl, time_bgl = db_connect.get_last_bgl(start_time)
    if bgl == "--" or bgl < 4 or bgl > 10:
        sheet.cell(row=row, column=5).font = font_red
    sheet.cell(row=row, column=5).value = bgl
    sheet.cell(row=row, column=3).value = time_bgl

    bgl, time_bgl = db_connect.get_last_bgl(start_time + 30 * 60)
    if bgl == "--" or bgl < 4 or bgl > 10:
        sheet.cell(row=row+1, column=5).font = font_red
    sheet.cell(row=row+1, column=5).value = bgl
    sheet.cell(row=row+1, column=3).value = time_bgl

    bgl, time_bgl = db_connect.get_last_bgl(start_time + 60 * 60)
    if bgl == "--" or bgl < 4 or bgl > 10:
        sheet.cell(row=row+2, column=5).font = font_red
    sheet.cell(row=row+2, column=5).value = bgl
    sheet.cell(row=row+2, column=3).value = time_bgl

    bgl, time_bgl = db_connect.get_last_bgl(start_time + 120 * 60)
    if bgl == "--" or bgl < 4 or bgl > 10:
        sheet.cell(row=row+3, column=5).font = font_red
    sheet.cell(row=row+3, column=5).value = bgl
    sheet.cell(row=row+3, column=3).value = time_bgl

    bgl, time_bgl = db_connect.get_last_bgl(start_time + 180 * 60)
    if bgl == "--" or bgl < 4 or bgl > 10:
        sheet.cell(row=row+4, column=5).font = font_red
    sheet.cell(row=row+4, column=5).value = bgl
    sheet.cell(row=row+4, column=3).value = time_bgl

    bgl, time_bgl = db_connect.get_last_bgl(start_time + 240 * 60)
    if bgl == "--" or bgl < 4 or bgl > 10:
        sheet.cell(row=row+5, column=5).font = font_red
    sheet.cell(row=row+5, column=5).value = bgl
    sheet.cell(row=row+5, column=3).value = time_bgl


def insert_data_sheet(sheet, row, bolus, bolus_type, carb_ratio, carb_input_xe, active_insulin, font, comments):
    # Показатели
    sheet.cell(row=row, column=7).value = bolus
    sheet.cell(row=row + 1, column=7).value = bolus_type
    sheet.cell(row=row + 2, column=7).value = carb_ratio
    sheet.cell(row=row + 3, column=7).value = carb_input_xe
    sheet.cell(row=row + 5, column=7).value = active_insulin
    sheet.cell(row=row, column=8).font = font
    sheet.cell(row=row, column=8).value = comments


def insert_data_sheet_addition(sheet, row, comments):
    tmp = sheet.cell(row=row+2, column=8).value
    if tmp is None:
        sheet.cell(row=row+2, column=8).value = comments
    else:
        sheet.cell(row=row+2, column=8).value = "{}\n{}".format(tmp, comments)


def insert_to_sheet(sheet, food_estimate, bolus, bolus_type, carb_ratio, carb_input_xe, active_insulin, font,
                    comments, time_of_day, addition, time_unix, db_connect):

    global breakfast
    global lunch
    global dinner

    if time_of_day == "breakfast":
        row = 7
        if breakfast == 0 and food_estimate != 0:
            insert_bgl(row, time_unix, db_connect, sheet)
            insert_data_sheet(sheet, row, bolus, bolus_type, carb_ratio, carb_input_xe, active_insulin,
                              font, comments)
            breakfast = breakfast + 1
        else:
            insert_data_sheet_addition(sheet, row, addition)

    elif time_of_day == "lunch":
        row = 14
        if lunch == 0 and food_estimate != 0:
            insert_bgl(row, time_unix, db_connect, sheet)
            insert_data_sheet(sheet, row, bolus, bolus_type, carb_ratio, carb_input_xe, active_insulin,
                              font, comments)
            lunch = lunch + 1
        else:
            insert_data_sheet_addition(sheet, row, addition)

    elif time_of_day == "dinner":
        row = 21
        if dinner == 0 and food_estimate != 0:
            insert_bgl(row, time_unix, db_connect, sheet)
            insert_data_sheet(sheet, row, bolus, bolus_type, carb_ratio, carb_input_xe, active_insulin,
                              font, comments)
            dinner = dinner + 1
        else:
            insert_data_sheet_addition(sheet, row, addition)


def fill_diary(date_ago):

    today = dt.datetime.combine(dt.datetime.today(), dt.time.min)
    to_tmp = today - dt.timedelta(days=date_ago)
    from_tmp = to_tmp - dt.timedelta(days=1)

    from_utc = int(time.mktime((from_tmp).timetuple()))
    to_utc = int(time.mktime((to_tmp).timetuple()))

    wb = openpyxl.load_workbook("blank.xlsx")
    sheet = wb["Лист1"]
    locale.setlocale(locale.LC_ALL, ('ru_RU', 'UTF-8'))
    sheet["A1"] = time.strftime("%d %B %Y", time.localtime(from_utc))
    sheet["F1"] = time.strftime("%A", time.localtime(from_utc))

    font_red = Font(name='Calibri', size=14, bold=True, color='FF0000')
    font_black = Font(name='Calibri', size=14, bold=True, color='000000')

    db_connect = db.DB()
    out = db_connect.get_bolus_wizard(from_utc, to_utc)
    dual_start_tmp = 0

    for next_data in out:
        time_of_day = next_data["time_of_day"]
        time_unix = next_data["time_unix"]

        print(next_data)

        if "dual" in next_data["type"]:
            if next_data["bolus_part"] == 2:
                normal = next_data["normal_programmed_amount"]
                square = next_data["square_programmed_amount"]
                total = next_data["total_programmed_amount"]

                normal_perc = int(round((normal * 100) / total, 0))
                square_perc = int(round((square * 100) / total, 0))

                # Показатели
                bolus = "{} ({}/{})".format(total, normal, square)
                bolus_type = "Д:{}/{} ({})".format(normal_perc, square_perc, next_data["programmed_duration"])
                carb_input_xe = "{} ({})".format(next_data["carb_input"] * 10, next_data["carb_input"])

                # Комменты
                dual_end_tmp = next_data["event_time"]
                dual_end = dual_end_tmp.split(" ")[1].split(":")

                dual_start = dual_start_tmp.split(" ")[1].split(":")
                bolus_start = "{}:{} ~ ".format(dual_start[0], dual_start[1])

                if next_data["canceled"] == 1:
                    comments = "{}{}:{}, {} ({}/{})".format(bolus_start, dual_end[0], dual_end[1],
                                                            next_data["delivered_duration"],
                                                            next_data["square_programmed_amount"],
                                                            next_data["delivered_amount"])

                    font = font_red
                else:
                    comments = "{}{}:{}".format(bolus_start, dual_end[0], dual_end[1])
                    font = font_black

                addition = "Д:{}, Б:{}, У(ХЕ):{}, К:{}, А:{}".format(comments, total, carb_input_xe,
                                                                     next_data["carb_ratio"],
                                                                     next_data["active_insulin"])

                insert_to_sheet(sheet, next_data["food_estimate"], bolus, bolus_type, next_data["carb_ratio"],
                                carb_input_xe, next_data["active_insulin"], font, comments, time_of_day, addition,
                                time_unix, db_connect)

            if next_data["bolus_part"] == 1:
                dual_start_tmp = next_data["event_time"]
            else:
                dual_start_tmp = 0

        if "normal" in next_data["type"]:
            # Показатели
            bolus = next_data["final_estimate"]

            bolus_type = "H:{}".format(next_data["final_estimate"])
            carb_input_xe = "{} ({})".format(next_data["carb_input"] * 10, next_data["carb_input"])
            normal_start = next_data["event_time"].split(" ")[1].split(":")
            bolus_start = "{}:{}".format(normal_start[0], normal_start[1])

            if next_data["canceled"] == 1:
                comments = "{} ({}/{})".format(bolus_start, next_data["programmed_amount"],
                                               next_data["delivered_amount"])
                font = font_red
            else:
                font = font_black
                comments = bolus_start

            addition = "Н:{}, Б:{}, У(ХЕ):{}, К:{}, А:{}".format(bolus_start, bolus, carb_input_xe,
                                                                 next_data["carb_ratio"], next_data["active_insulin"])

            insert_to_sheet(sheet, next_data["food_estimate"], bolus, bolus_type, next_data["carb_ratio"],
                            carb_input_xe, next_data["active_insulin"], font, comments, time_of_day, addition,
                            time_unix, db_connect)

        if "square" in next_data["type"]:
            # Показатели
            bolus = next_data["final_estimate"]
            bolus_type = "К:{}".format(next_data["final_estimate"])
            carb_input_xe = "{} ({})".format(next_data["carb_input"] * 10, next_data["carb_input"])

            normal_start = next_data["event_time"].split(" ")[1].split(":")
            bolus_start = "{}:{},".format(normal_start[0], normal_start[1])

            if next_data["canceled"] == 1:
                comments = "{} {} ({}/{})".format(bolus_start, next_data["delivered_duration"],
                                                  next_data["programmed_amount"], next_data["delivered_amount"])
                font = font_red
            else:
                font = font_black
                comments = bolus_start

            addition = "К:{}, Б:{}, У(ХЕ):{}, К:{}, А:{}".format(comments, bolus, carb_input_xe,
                                                                 next_data["carb_ratio"], next_data["active_insulin"])

            insert_to_sheet(sheet, next_data["food_estimate"], bolus, bolus_type, next_data["carb_ratio"],
                            carb_input_xe, next_data["active_insulin"], font, comments, time_of_day, addition,
                            time_unix, db_connect)

    insert_bgl_night(from_utc, db_connect, sheet)

    next_data = db_connect.get_history_daily_totals(from_utc, to_utc + 60 * 60)
    # print(next_data)

    sheet.cell(row=3, column=6).value = "{} ({}%)".format(next_data["basal_insulin"], next_data["basal_percent"])
    sheet.cell(row=3, column=7).value = "{} ({}%)".format(next_data["bolus_insulin"], next_data["bolus_percent"])
    sheet.cell(row=3, column=8).value = "{}".format(next_data["total_insulin"])

    sheet.cell(row=33, column=7).value = "{}".format(next_data["total_bolus_wizard_insulin_as_food_only_bolus"])
    sheet.cell(row=34, column=7).value = "{} ({})".format(next_data["total_food_input"] * 10,
                                                          next_data["total_food_input"])

    sheet.cell(row=6, column=1).value = "Измерений в день:{}, Сахар(Средний):{}".format(
        next_data["meter_bg_count"], round((next_data["sg_average"] / cnl24lib.NGPConstants.BG_UNITS.MMOLXLFACTOR), 1))

    wb.save("report.xlsx")
    wb.close()
